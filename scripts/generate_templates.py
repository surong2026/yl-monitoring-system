"""
生成监测数据 Excel 模板文件

两种格式:
- column_format.xlsx — 列格式 (宽表), 每行一个断面
- row_format.xlsx — 行格式 (长表), 每行一个参数
"""

import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from datetime import date
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

TEMPLATE_DIR = Path(__file__).resolve().parent.parent / "data" / "templates"

HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
BODY_FONT = Font(name="微软雅黑", size=10)
EXAMPLE_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
THIN_BORDER = Side(style="thin", color="B4B4B4")


def _style_header(ws, row: int, col_count: int):
    """样式化表头"""
    border = Border(left=THIN_BORDER, right=THIN_BORDER, top=THIN_BORDER, bottom=THIN_BORDER)
    for c in range(1, col_count + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border


def _style_body(ws, row: int, col_count: int, is_example: bool = False):
    """样式化数据行"""
    border = Border(left=THIN_BORDER, right=THIN_BORDER, top=THIN_BORDER, bottom=THIN_BORDER)
    for c in range(1, col_count + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = BODY_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
        if is_example:
            cell.fill = EXAMPLE_FILL


def generate_column_template():
    """生成列格式模板 (宽表)"""
    wb = Workbook()
    ws = wb.active
    ws.title = "地表水监测数据"

    headers = ["断面编码", "断面名称", "采样日期",
               "pH", "DO", "COD", "CODMn", "BOD5", "NH3N", "TP", "TN",
               "Cu", "Zn", "F", "As", "Hg", "Cd", "Cr6", "Pb",
               "CN", "VPH", "LAS", "S2", "FC"]
    unit_row = ["", "", "",
                "无量纲", "mg/L", "mg/L", "mg/L", "mg/L", "mg/L", "mg/L", "mg/L",
                "mg/L", "mg/L", "mg/L", "mg/L", "mg/L", "mg/L", "mg/L", "mg/L",
                "mg/L", "mg/L", "mg/L", "mg/L", "个/L"]

    for i, h in enumerate(headers, 1):
        ws.cell(row=1, column=i, value=h)
    _style_header(ws, 1, len(headers))

    for i, u in enumerate(unit_row, 1):
        ws.cell(row=2, column=i, value=u)
    _style_body(ws, 2, len(unit_row))

    example = ["YL-WS-001", "南流江六司段", "2026-01-15",
               7.3, 7.1, 12.0, 3.5, 2.8, 0.35, 0.08, 1.2,
               0.005, 0.02, 0.5, 0.002, 0.00003, 0.001, 0.02, 0.01,
               0.005, 0.002, 0.1, 0.05, 5000]
    for i, v in enumerate(example, 1):
        ws.cell(row=3, column=i, value=v)
    _style_body(ws, 3, len(example), is_example=True)

    # 列宽
    widths = [15, 18, 14] + [10] * (len(headers) - 3)
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    path = TEMPLATE_DIR / "column_format.xlsx"
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    print(f"列格式模板已生成: {path}")


def generate_row_template():
    """生成行格式模板 (长表)"""
    wb = Workbook()
    ws = wb.active
    ws.title = "地表水监测数据"

    headers = ["断面编码", "断面名称", "采样日期", "监测项目", "单位", "监测值", "检出限", "备注"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=1, column=i, value=h)
    _style_header(ws, 1, len(headers))

    example_data = [
        ["YL-WS-001", "南流江六司段", "2026-01-15", "pH", "无量纲", 7.3, "", ""],
        ["YL-WS-001", "南流江六司段", "2026-01-15", "溶解氧", "mg/L", 7.1, "", ""],
        ["YL-WS-001", "南流江六司段", "2026-01-15", "化学需氧量", "mg/L", 12.0, "", ""],
        ["YL-WS-001", "南流江六司段", "2026-01-15", "氨氮", "mg/L", 0.35, "", ""],
        ["YL-WS-001", "南流江六司段", "2026-01-15", "总磷", "mg/L", 0.08, "", ""],
        ["YL-WS-001", "南流江六司段", "2026-01-15", "汞", "mg/L", 0.00003, 0.00004, "低于检出限"],
    ]
    for r, row_data in enumerate(example_data, 2):
        for c, v in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=v)
        _style_body(ws, r, len(headers), is_example=True)

    widths = [15, 18, 14, 18, 10, 12, 10, 20]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    path = TEMPLATE_DIR / "row_format.xlsx"
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    print(f"行格式模板已生成: {path}")


if __name__ == "__main__":
    generate_column_template()
    generate_row_template()
